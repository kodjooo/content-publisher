from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config.models import ExcelConfig
from app.crawler.models import ProductSnapshot, SiteCrawlResult
from app.crawler.site_crawler import format_price_for_excel
from app.crawler.utils import normalize_url
from app.logger import get_logger

logger = get_logger(__name__)


@dataclass(slots=True)
class ExcelUpdateSummary:
    updated: int = 0
    skipped_not_found: int = 0
    processed: int = 0


class ExcelUpdater:
    """Обновляет существующий Excel-файл по совпадению URL товара."""

    def __init__(self, config: ExcelConfig, strip_params: list[str]):
        self.config = config
        self.strip_params = strip_params

    def apply(self, results: list[SiteCrawlResult]) -> ExcelUpdateSummary:
        workbook_path = Path(self.config.workbook_path)
        if not workbook_path.exists():
            raise FileNotFoundError(f"Excel-файл не найден: {workbook_path}")

        wb = load_workbook(workbook_path)
        try:
            ws = self._resolve_sheet(wb)
            col_map = self._resolve_columns(ws)
            url_col = self._resolve_url_column(col_map)
            price_col = self._ensure_column(ws, col_map, self.config.price_column_name)
            rating_col = self._ensure_column(ws, col_map, self.config.rating_column_name)
            in_stock_col = self._ensure_column(ws, col_map, self.config.in_stock_column_name)

            row_index_by_url = self._build_url_index(ws, url_col)
            summary = ExcelUpdateSummary()
            for site_result in results:
                for item in site_result.records:
                    summary.processed += 1
                    row_idx = row_index_by_url.get(self._norm(item.product_url))
                    if row_idx is None:
                        summary.skipped_not_found += 1
                        continue
                    ws.cell(row=row_idx, column=price_col).value = format_price_for_excel(item.current_price)
                    ws.cell(row=row_idx, column=rating_col).value = item.rating
                    ws.cell(row=row_idx, column=in_stock_col).value = item.in_stock
                    summary.updated += 1
            wb.save(workbook_path)
            return summary
        finally:
            wb.close()

    def _resolve_sheet(self, wb) -> Worksheet:
        if self.config.sheet_name:
            if self.config.sheet_name not in wb.sheetnames:
                raise ValueError(f"Лист '{self.config.sheet_name}' не найден в Excel")
            return wb[self.config.sheet_name]
        return wb.active

    def _resolve_columns(self, ws: Worksheet) -> dict[str, int]:
        header_row = self.config.header_row
        columns: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=header_row, column=col).value
            if value is None:
                continue
            key = str(value).strip().lower()
            if not key:
                continue
            columns[key] = col
        return columns

    def _resolve_url_column(self, col_map: dict[str, int]) -> int:
        for candidate in self.config.url_column_candidates:
            key = candidate.strip().lower()
            if key in col_map:
                return col_map[key]
        raise ValueError(
            "Не найдена колонка ссылки в Excel. Проверьте EXCEL_URL_COLUMN_CANDIDATES."
        )

    def _ensure_column(self, ws: Worksheet, col_map: dict[str, int], header: str) -> int:
        key = header.strip().lower()
        existing = col_map.get(key)
        if existing is not None:
            return existing
        new_col = ws.max_column + 1
        ws.cell(row=self.config.header_row, column=new_col).value = header
        col_map[key] = new_col
        return new_col

    def _build_url_index(self, ws: Worksheet, url_col: int) -> dict[str, int]:
        index: dict[str, int] = {}
        start = self.config.header_row + 1
        for row_idx in range(start, ws.max_row + 1):
            raw = ws.cell(row=row_idx, column=url_col).value
            if raw is None:
                continue
            raw_str = str(raw).strip()
            if not raw_str:
                continue
            index[self._norm(raw_str)] = row_idx
        return index

    def _norm(self, url: str) -> str:
        normalized, _ = normalize_url(url, base_url=None, strip_params=self.strip_params)
        return normalized
