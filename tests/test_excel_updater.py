from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.config.models import ExcelConfig
from app.crawler.models import ProductSnapshot, SiteCrawlResult
from app.excel import ExcelUpdater


def _create_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["product_url", "name"])
    ws.append(["https://demo.example/p/1", "A"])
    ws.append(["https://demo.example/p/2", "B"])
    wb.save(path)
    wb.close()


def test_excel_updater_updates_rows_by_url(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "products.xlsx"
    _create_workbook(xlsx_path)

    updater = ExcelUpdater(
        ExcelConfig(
            workbook_path=xlsx_path,
            sheet_name="Products",
            url_column_candidates=["product_url"],
            price_column_name="current_price",
            rating_column_name="rating",
            in_stock_column_name="in_stock",
        ),
        strip_params=["utm_*"]
    )

    results = [
        SiteCrawlResult(
            site_name="demo",
            records=[
                ProductSnapshot(
                    source_site="demo.example",
                    category_url="https://demo.example/cat",
                    product_url="https://demo.example/p/1?utm_source=t",
                    page_num=1,
                    current_price=1500.0,
                    rating=4.7,
                    in_stock=True,
                ),
                ProductSnapshot(
                    source_site="demo.example",
                    category_url="https://demo.example/cat",
                    product_url="https://demo.example/p/3",
                    page_num=1,
                    current_price=900.0,
                    rating=4.1,
                    in_stock=False,
                ),
            ],
            metrics=[],
        )
    ]

    summary = updater.apply(results)

    assert summary.processed == 2
    assert summary.updated == 1
    assert summary.skipped_not_found == 1

    wb = load_workbook(xlsx_path)
    ws = wb["Products"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    price_col = headers.index("current_price") + 1
    rating_col = headers.index("rating") + 1
    stock_col = headers.index("in_stock") + 1
    assert ws.cell(row=2, column=price_col).value == 1500
    assert ws.cell(row=2, column=rating_col).value == 4.7
    assert ws.cell(row=2, column=stock_col).value is True
    wb.close()
